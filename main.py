from fastapi import FastAPI, Depends, HTTPException
from sqlalchemy.orm import Session
from database import SessionLocal, engine
from models import Base, Boleto, Observacao
from schemas import BoletoCreate, BoletoResponse, ObservacaoCreate, ObservacaoResponse
from typing import List
from fastapi.middleware.cors import CORSMiddleware
from datetime import timedelta
from fastapi import UploadFile, File
import csv
import io
from openpyxl import load_workbook
from fastapi.responses import FileResponse
from datetime import datetime
import shutil

Base.metadata.create_all(bind=engine)

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@app.post("/boletos/importar/")
def importar_boletos(file: UploadFile = File(...), db: Session = Depends(get_db)):
    filename = file.filename.lower()
    boletos = []

    if filename.endswith(".csv"):
        content = file.file.read().decode("utf-8")
        reader = csv.DictReader(io.StringIO(content))
        for row in reader:
            boletos.append(row)

    elif filename.endswith(".xlsx"):
        wb = load_workbook(file.file, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            boletos.append(dict(zip(headers, row)))
    else:
        raise HTTPException(status_code=400, detail="Formato de arquivo não suportado (use .csv ou .xlsx)")

    for row in boletos:
        try:
            boleto = Boleto(
                empresa=row["empresa"],
                devedor=row["devedor"],
                localidade=row["localidade"],
                valor_parcela=float(row["valor_parcela"]),
                valor_total=float(row["valor_total"]),
                qtd_parcelas=int(row["qtd_parcelas"]),
                parcela_atual=int(row["parcela_atual"]),
                data_geracao=datetime.strptime(str(row["data_geracao"]), "%Y-%m-%d").date(),
                data_vencimento=datetime.strptime(str(row["data_vencimento"]), "%Y-%m-%d").date(),
                status=row["status"]
            )
            db.add(boleto)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Erro ao importar linha: {row} — {e}")

    db.commit()
    return {"importados": len(boletos)}

@app.post("/boletos/", response_model=BoletoResponse)
def create_boleto(boleto: BoletoCreate, db: Session = Depends(get_db)):
    db_boleto = Boleto(**boleto.dict())
    db.add(db_boleto)
    db.commit()
    db.refresh(db_boleto)
    return db_boleto

@app.get("/boletos/", response_model=List[BoletoResponse])
def listar_boletos(
    empresa: str = "",
    devedor: str = "",
    localidade: str = "",
    status: str = "",
    db: Session = Depends(get_db)
):
    query = db.query(Boleto)
    if empresa:
        query = query.filter(Boleto.empresa.ilike(f"%{empresa}%"))
    if devedor:
        query = query.filter(Boleto.devedor.ilike(f"%{devedor}%"))
    if localidade:
        query = query.filter(Boleto.localidade.ilike(f"%{localidade}%"))
    if status:
        query = query.filter(Boleto.status.ilike(f"%{status}%"))
    return query.all()


@app.put("/boletos/{boleto_id}", response_model=BoletoResponse)
def editar_boleto(boleto_id: int, dados: BoletoCreate, db: Session = Depends(get_db)):
    boleto = db.query(Boleto).get(boleto_id)
    if not boleto:
        raise HTTPException(status_code=404, detail="Boleto não encontrado")
    for key, value in dados.dict().items():
        setattr(boleto, key, value)
    db.commit()
    db.refresh(boleto)
    return boleto

@app.delete("/boletos/{boleto_id}")
def excluir_boleto(boleto_id: int, db: Session = Depends(get_db)):
    boleto = db.query(Boleto).get(boleto_id)
    if not boleto:
        raise HTTPException(status_code=404, detail="Boleto não encontrado")
    db.delete(boleto)
    db.commit()
    return {"mensagem": "Boleto excluído"}

@app.post("/boletos/{boleto_id}/cobrado", response_model=BoletoResponse)
def marcar_cobrado(boleto_id: int, db: Session = Depends(get_db)):
    boleto = db.query(Boleto).get(boleto_id)
    if not boleto:
        raise HTTPException(status_code=404, detail="Boleto não encontrado")

    if boleto.parcela_atual < boleto.qtd_parcelas:
        boleto.parcela_atual += 1
        boleto.data_vencimento = boleto.data_vencimento + timedelta(days=30)
    else:
        boleto.status = "cobrada"

    db.commit()
    db.refresh(boleto)
    return boleto

@app.post("/boletos/{boleto_id}/observacoes/", response_model=ObservacaoResponse)
def adicionar_observacao(boleto_id: int, obs: ObservacaoCreate, db: Session = Depends(get_db)):
    boleto = db.query(Boleto).get(boleto_id)
    if not boleto:
        raise HTTPException(status_code=404, detail="Boleto não encontrado")
    nova_obs = Observacao(**obs.dict(), boleto_id=boleto_id)
    db.add(nova_obs)
    db.commit()
    db.refresh(nova_obs)
    return nova_obs

@app.get("/boletos/{boleto_id}/observacoes/", response_model=List[ObservacaoResponse])
def listar_observacoes(boleto_id: int, db: Session = Depends(get_db)):
    return db.query(Observacao).filter(Observacao.boleto_id == boleto_id).all()

@app.put("/observacoes/{obs_id}", response_model=ObservacaoResponse)
def editar_observacao(obs_id: int, obs: ObservacaoCreate, db: Session = Depends(get_db)):
    observacao = db.query(Observacao).get(obs_id)
    if not observacao:
        raise HTTPException(status_code=404, detail="Observação não encontrada")
    observacao.data = obs.data
    observacao.descricao = obs.descricao
    db.commit()
    db.refresh(observacao)
    return observacao

@app.delete("/observacoes/{obs_id}")
def excluir_observacao(obs_id: int, db: Session = Depends(get_db)):
    observacao = db.query(Observacao).get(obs_id)
    if not observacao:
        raise HTTPException(status_code=404, detail="Observação não encontrada")
    db.delete(observacao)
    db.commit()
    return {"mensagem": "Observação excluída"}

@app.get("/exportar-db")
def exportar_db():
    return FileResponse("boletos.db", filename="boletos_backup.db", media_type='application/octet-stream')

@app.get("/criar-backup")
def criar_backup():
    agora = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_backup = f"backup_{agora}.db"
    shutil.copy("boletos.db", nome_backup)
    return {"mensagem": f"Backup criado: {nome_backup}"}

   
@app.post("/restaurar-backup-upload")
def restaurar_backup_upload(file: UploadFile = File(...)):
    try:
        with open("boletos.db", "wb") as f:
            f.write(file.file.read())
        return {"mensagem": "Backup restaurado com sucesso"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
import os

# Monta arquivos estáticos se houver
if os.path.isdir("static"):
    app.mount("/static", StaticFiles(directory="static"), name="static")

# Rota para servir o index.html
@app.get("/", response_class=HTMLResponse)
def read_index():
    with open("index.html", "r", encoding="utf-8") as f:
        return f.read()